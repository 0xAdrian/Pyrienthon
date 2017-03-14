#!/usr/bin/python
# Pyrienthon v.Beta
# @AdTrejo

from os import listdir
import xml.etree.ElementTree as ET
from openpyxl import Workbook
import xlrd

wb = Workbook()
index = 2

def banner():
	print "           _________"
	print "         /'        /|"
	print "        /         / |_"
	print "       /         /  //|"
	print "      /_________/  ////|"
	print "     |   _ _    | 8o////|"
	print "     | /'// )_  |   8///|"
	print "     |/ // // ) |   8o///|"
	print "     / // // //,|  /  8//|"
	print "    / // // /// | /   8//|"
	print "   / // // ///__|/    8//|"
	print "  /.(_)// /// |       8///|"
	print " (_)' `(_)//| |       8////|___________"
	print "(_) /_\ (_)'| |        8///////////////"
	print "(_) \ / (_)'|_|         8/////////////"
	print " (_)._.(_) d' Hb         8oooooooopb'"
	print "   `(_)'  d'  H`b"
	print "         d'   `b`b"
	print "        d'     H `b"
	print "       d'      `b `b"
	print "      d'           `b"
	print "     d'             `b"
	return

def complete(pluginID):
	ind = str(index)
	vuln_db = xlrd.open_workbook('vulnerabilidades.xls')
	worksheet = vuln_db.sheet_by_index(0)
	id_col = 1
	tit_col = 6
	desc_col = 7
	sol_col = 8
	curr_row = 1

	while curr_row < worksheet.nrows:
		id_value = worksheet.cell_value(curr_row,id_col)
		try:
			nessus_id = int(id_value)
		except:
			nessus_id = 0

		if nessus_id == int(pluginID):
			titulo =  worksheet.cell_value(curr_row,tit_col)
			descripcion =  worksheet.cell_value(curr_row,desc_col)
			solucion =  worksheet.cell_value(curr_row,sol_col)
			ws = wb.get_active_sheet()
			ws['A'+ind] = worksheet.cell_value(curr_row,tit_col)
			ws['H'+ind] = worksheet.cell_value(curr_row,desc_col)
			ws['I'+ind] = worksheet.cell_value(curr_row,sol_col)
			wb.save('Reporte.xlsx')
			break
		curr_row = curr_row + 1
	return

def pop(pluginName, pluginID, cve, cvss, severity, host, port):
	ws = wb.get_active_sheet()
	ind = str(index)
	ws['A'+ind] = pluginName
	ws['B'+ind] = int(pluginID)
	ws['C'+ind] = cve
	ws['D'+ind] = cvss
	ws['E'+ind] = severity
	ws['F'+ind] = host
	ws['G'+ind] = int(port)
	wb.save('Reporte.xlsx')
	return

def initExcel():
	ws = wb.get_active_sheet()
	ws.title = 'Vulnerabilidades'
	ws['A1'] = 'Nombre de la vulnerabilidad'
	ws['B1'] = 'Nessus ID'
	ws['C1'] = 'CVE'
	ws['D1'] = 'CVSS'
	ws['E1'] = 'Prioridad'
	ws['F1'] = 'Direccion IP'
	ws['G1'] = 'Puerto'
	ws['H1'] = 'Descripcion'
	ws['I1'] = 'Solucion'
	wb.save('Reporte.xlsx')
	return

def process(file_name):
	print 'Processing ' + file_name +'...'
	tree = ET.parse('./Nessus/'+file_name)
	root = tree.getroot()
	global index

	for rHost in root.iter('ReportHost'):
		host = rHost.get('name')
		for rItem in rHost.iter('ReportItem'):
			severity = int(rItem.get('severity'))
			if severity > 0:
				pluginName = rItem.get('pluginName')
				pluginID = rItem.get('pluginID')
				cve = ''
				for cve_item in rItem.iter('cve'):
				 	cve = cve + cve_item.text +'\n'
				if cve == '':
				 	cve = 'N/A'
				try:
					cbs = rItem.find('cvss_base_score').text
					cv = rItem.find('cvss_vector').text
					cvss = 'CVSS Base Score: ' + rItem.find('cvss_base_score').text + ' (' + rItem.find('cvss_vector').text + ')'
				except :
					cvss = 'N/A'
				
				port = rItem.get('port')
				pop(pluginName, pluginID, cve, cvss, severity, host, port)
				complete(pluginID)
				index = index+1
	return

files = listdir('./Nessus')
banner()
initExcel()
for f in files:
	process(f)
